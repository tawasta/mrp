import base64
import csv
import io
from io import BytesIO
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import logging

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

_logger = logging.getLogger(__name__)
# Required file columns. These must exist in the header row.
REQUIRED_COLS = {
    # Header/template identifiers
    "product_tmpl_id",
    "product_tmpl_name",
    "product_tmpl_uom_id",
    "product_qty",
    "code",
    "type",
    # Component line identifiers
    "line_product_id",
    "line_product_name",
    "line_product_qty",
    "line_product_uom",
    "line_product_uom_id",
}
ALLOWED_TYPES = {"normal", "phantom"}


class MrpBomImportWizard(models.TransientModel):
    """
    BoM Import Wizard (CSV/XLSX)

    Purpose:
        Import Bills of Materials from CSV/XLSX where each spreadsheet row represents
        a single BoM component line. Lines are grouped into BoMs by:
        (product_tmpl_id, code, type, product_qty).

    Usage:
        Open the wizard, upload a file (CSV or XLSX), optionally run Dry Run to validate,
        then import. The wizard can create missing products if enabled.

    Extensibility:
        Child modules can add new import columns, extra validations, and creation values
        via a small set of hooks without overriding action_import.
    """

    _name = "mrp.bom.import.wizard"
    _description = "Import Bills of Materials (CSV/XLSX; one row per component)"

    data_file = fields.Binary(string="File", required=False)
    filename = fields.Char("Filename")
    create_missing_products = fields.Boolean(
        string="Create Missing Products", default=True
    )
    company_id = fields.Many2one(
        "res.company", string="Company", default=lambda s: s.env.company
    )
    dry_run = fields.Boolean(string="Dry run (validate only)", default=False)

    @api.model
    def extra_required_columns(self):
        """Return extra required column names added by extensions."""
        return set()

    @api.model
    def parse_extra_columns(self, raw_row, rownum):
        """
        Parse extension-specific columns from raw_row and return a dict to merge into row_norm.
        Extensions override this to add their own fields.
        """
        return {}

    def validate_extra_row(self, row_norm):
        """
        Add per-row extra validations. Raise UserError on failure.
        Extensions override this to enforce their rules.
        """
        return

    def mutate_bom_create_vals(self, vals, header_rows):
        """
        Adjust mrp.bom create values before creation. Return the (possibly) modified dict.
        Extensions override this to inject extra fields.
        """
        return vals

    def mutate_line_create_vals(self, vals, row_norm):
        """
        Adjust mrp.bom.line create values before creation. Return the (possibly) modified dict.
        Extensions override this to inject extra fields.
        """
        return vals

    # ---------- helpers ----------

    @staticmethod
    def _clean(s):
        """Return a trimmed string for a possibly None/empty value. Used by all parsers."""
        return (s or "").strip()

    def _ensure_columns(self, fieldnames):
        """Check that the uploaded file contains required columns (base + extension)."""
        required = set(REQUIRED_COLS) | set(self.extra_required_columns() or set())
        missing = required - set(fieldnames or [])
        if missing:
            raise UserError(
                _("Missing required columns: %s") % ", ".join(sorted(missing))
            )

    def _parse_int_id(self, value, field_label, rownum=None, allow_empty=True):
        """
        Parse a positive integer ID (e.g., uom_id) possibly written as '5' or '5.0'.
        Returns None if empty and allow_empty=True. Used for product_tmpl_uom_id / line_product_uom_id.
        """
        v = self._clean(value)
        if not v:
            return (
                None
                if allow_empty
                else self._raise(_("%s is required") % field_label, rownum)
            )
        try:
            iv = int(float(v))
        except Exception:
            where = (" (row %s)" % rownum) if rownum else ""
            raise UserError(_("Invalid %s '%s'%s") % (field_label, value, where))
        if iv <= 0:
            where = (" (row %s)" % rownum) if rownum else ""
            raise UserError(_("%s must be a positive ID%s") % (field_label, where))
        return iv

    def _raise(self, msg, rownum=None):
        """Raise a UserError with optional row-number context. Used across validators."""
        where = (" (row %s)" % rownum) if rownum else ""
        raise UserError(msg + where)

    def _get_product_template(self, tmpl_code, name=None, uom_id=None, standard_price=None):
        """
        Resolve or create product.template by default_code (SKU/code). Used for BoM header product.
        Enforces name is given and differs from code when creating.
        """
        ProductTemplate = self.env["product.template"].with_context(active_test=False)
        pt = ProductTemplate.search([("default_code", "=", tmpl_code)], limit=1)
        if pt:
            return pt
        if not self.create_missing_products:
            raise UserError(_("Product template '%s' not found.") % tmpl_code)
        if not name or name == tmpl_code:
            raise UserError(
                _("Name is required and must differ from code for template '%s'.")
                % tmpl_code
            )
        vals = {
            "name": name,
            "default_code": tmpl_code,
            "detailed_type": "product",
            "company_id": self.company_id.id,
            "standard_price": standard_price or 0.0,
        }
        if uom_id:
            vals.update({"uom_id": uom_id, "uom_po_id": uom_id})
        _logger.info("Creating missing product template: %s", vals)
        return ProductTemplate.create(vals)

    def _get_product(self, line_code, name=None, uom_id=None, standard_price=None):
        """
        Resolve or create product.product by default_code (code). Used for component lines.
        Enforces name is given and differs from code when creating.
        """
        ProductProduct = self.env["product.product"].with_context(active_test=False)
        p = ProductProduct.search([("default_code", "=", line_code)], limit=1)
        if p:
            return p
        if not self.create_missing_products:
            raise UserError(_("Product '%s' not found.") % line_code)
        if not name or name == line_code:
            raise UserError(
                _("Name is required and must differ from code for product '%s'.")
                % line_code
            )
        vals = {
            "name": name,
            "default_code": line_code,
            "detailed_type": "product",
            "company_id": self.company_id.id,
            "standard_price": standard_price or 0.0,
        }
        if uom_id:
            vals.update({"uom_id": uom_id, "uom_po_id": uom_id})
        _logger.info("Creating missing product: %s", vals)
        return ProductProduct.create(vals)

    def _get_uom(self, name):
        """
        Find a UoM by name (exact first, then ilike). Used to set BoM line UoM when given by name.
        Returns a uom.uom recordset. Name must be provided (required).
        """
        if not name:
            raise UserError(_("Unit of Measure name is required"))
        Uom = self.env["uom.uom"]
        uom = Uom.search([("name", "=", name)], limit=1) or Uom.search(
            [("name", "ilike", name)], limit=1
        )
        if not uom:
            raise UserError(_("Unit of Measure '%s' not found.") % name)
        return uom

    def _get_uom_by_id(self, id_):
        """
        Validate a UoM by ID. Used for product_tmpl_uom_id and line_product_uom_id to ensure existence.
        Returns the uom.uom recordset.
        """
        if not id_:
            raise UserError(_("Unit of Measure ID is required"))
        Uom = self.env["uom.uom"]
        uom = Uom.browse(id_)
        if not uom.exists():
            raise UserError(_("Unit of Measure with ID %s not found.") % id_)
        return uom

    def _parse_float(self, value, field_label, rownum=None, allow_zero=False):
        """
        Parse a float quantity. Used for product_qty and line_product_qty.
        Enforces positive values (allows zero for component lines when allow_zero=True).
        """
        try:
            val = float(value)
        except Exception:
            where = (" (row %s)" % rownum) if rownum else ""
            raise UserError(_("Invalid %s '%s'%s") % (field_label, value, where))
        if val < 0 or (not allow_zero and val == 0):
            where = (" (row %s)" % rownum) if rownum else ""
            raise UserError(_("%s must be positive%s") % (field_label, where))
        return val

    # ---------- readers ----------

    def _read_csv_rows(self, content: bytes):
        """Stream CSV rows as dicts and validate headers. Used when the uploaded file is .csv."""
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        self._ensure_columns(reader.fieldnames)
        for row in reader:
            yield row

    def _read_xlsx_rows(self, content: bytes):
        """
        Stream XLSX rows as dicts and validate headers. Used when the uploaded file is .xlsx/.xlsm.
        Uses sheet 'BoM Import' if present, otherwise the active sheet.
        """
        if not load_workbook:
            raise UserError(_("XLSX reading is not available (missing 'openpyxl')."))
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb["BoM Import"] if "BoM Import" in wb.sheetnames else wb.active
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [("" if v is None else str(v)).strip() for v in row]
            if i == 1:
                headers = values
                self._ensure_columns(headers)
                continue
            yield dict(zip(headers, values))

    # ---------- action ----------

    def action_import(self):
        """
        Button entry: decode file, parse rows, validate, prevent duplicates, and create BoMs/lines.
        In Dry Run mode, only validates and shows a notification.
        """
        if not self.data_file or not self.filename:
            raise UserError(_("Please select a file first."))

        ext = self.filename.lower().rsplit(".", 1)[-1]
        content = base64.b64decode(self.data_file)

        if ext == "csv":
            raw_rows = list(self._read_csv_rows(content))
        elif ext in ("xlsx", "xlsm"):
            raw_rows = list(self._read_xlsx_rows(content))
        else:
            raise UserError(_("Unsupported file type: %s (use CSV or XLSX).") % ext)

        # Pre-validate & group
        groups = {}  # key -> [rows]
        errors = []
        all_rows = []

        for idx, row in enumerate(raw_rows, start=2):  # header at row 1
            # --- Header (BoM owner template) ---
            # product_tmpl_id == product.template.default_code (code)
            tmpl_code = self._clean(row.get("product_tmpl_id"))
            tmpl_name = self._clean(row.get("product_tmpl_name"))
            tmpl_uom_id_raw = self._clean(row.get("product_tmpl_uom_id"))

            code = self._clean(row.get("code"))
            rtype = (self._clean(row.get("type")) or "normal").lower()
            pqty_raw = self._clean(row.get("product_qty"))

            # --- Component line product ---
            # line_product_id == product.product.default_code (code)
            line_code = self._clean(row.get("line_product_id"))
            line_name = self._clean(row.get("line_product_name"))
            lqty_raw = self._clean(row.get("line_product_qty"))
            luom_name = self._clean(row.get("line_product_uom"))  # required by name
            line_uom_id_raw = self._clean(row.get("line_product_uom_id"))  # required

            standard_price_raw = self._clean(row.get("standard_price"))

            # Basic presence / format checks
            if not tmpl_code:
                errors.append(_("Row %s: 'product_tmpl_id' is empty.") % idx)
            if not tmpl_name:
                errors.append(_("Row %s: 'product_tmpl_name' is empty.") % idx)
            if tmpl_name and tmpl_code and tmpl_name == tmpl_code:
                errors.append(
                    _(
                        "Row %s: product_tmpl_name must differ from product_tmpl_id (code)."
                    )
                    % idx
                )

            if not code:
                errors.append(
                    _("Row %s: 'code' is empty. Provide a unique BoM code.") % idx
                )
            if rtype not in ALLOWED_TYPES:
                errors.append(_("Row %s: 'type' must be 'normal' or 'phantom'.") % idx)

            if not line_code:
                errors.append(_("Row %s: 'line_product_id' is empty.") % idx)
            if not line_name:
                errors.append(_("Row %s: 'line_product_name' is empty.") % idx)
            if line_name and line_code and line_name == line_code:
                errors.append(
                    _(
                        "Row %s: line_product_name must differ from line_product_id (code)."
                    )
                    % idx
                )
            if not luom_name:
                errors.append(
                    _("Row %s: 'line_product_uom' is empty (name required).") % idx
                )

            try:
                header_qty = self._parse_float(pqty_raw, "product_qty", idx)
            except UserError as e:
                errors.append(e.args[0] if e.args else str(e))
                header_qty = None
            try:
                line_qty = self._parse_float(
                    lqty_raw, "line_product_qty", idx, allow_zero=True
                )
            except UserError as e:
                errors.append(e.args[0] if e.args else str(e))
                line_qty = None

            # uom_id ints (required)
            try:
                header_uom_id = self._parse_int_id(
                    tmpl_uom_id_raw, "product_tmpl_uom_id", idx, allow_empty=False
                )
            except UserError as e:
                errors.append(e.args[0] if e.args else str(e))
                header_uom_id = None

            try:
                line_prod_uom_id = self._parse_int_id(
                    line_uom_id_raw, "line_product_uom_id", idx, allow_empty=False
                )
            except UserError as e:
                errors.append(e.args[0] if e.args else str(e))
                line_prod_uom_id = None

            standard_price = None
            if standard_price_raw:
                try:
                    standard_price = self._parse_float(
                        standard_price_raw, "standard_price", idx, allow_zero=True
                    )
                except UserError as e:
                    errors.append(e.args[0] if e.args else str(e))
                    standard_price = None

            if (
                not tmpl_code
                or not tmpl_name
                or header_uom_id is None
                or not code
                or rtype not in ALLOWED_TYPES
                or header_qty is None
                or not line_code
                or not line_name
                or line_qty is None
                or not luom_name
                or line_prod_uom_id is None
            ):
                continue

            # BoM line UoM (by name) — required
            try:
                line_uom = self._get_uom(luom_name)
            except UserError as e:
                errors.append(_("Row %s: %s") % (idx, e.args[0] if e.args else str(e)))
                line_uom = None
            if line_uom is None:
                continue

            # Base normalized fields
            row_norm = {
                "rownum": idx,
                # header/template
                "tmpl_code": tmpl_code,
                "tmpl_name": tmpl_name,
                "header_uom_id": header_uom_id,  # int
                "code": code,
                "type": rtype,
                "header_qty": header_qty,
                # component/product
                "line_code": line_code,
                "line_name": line_name,
                "line_qty": line_qty,
                "line_uom": line_uom,  # record
                "line_product_uom_id": line_prod_uom_id,  # int
                "standard_price": standard_price,
            }

            # Merge extension-specific parsed fields
            try:
                extra = self.parse_extra_columns(row, idx) or {}
                if extra:
                    row_norm.update(extra)
            except UserError as e:
                errors.append(e.args[0] if e.args else str(e))
                continue

            # Extra per-row validation from extensions
            try:
                self.validate_extra_row(row_norm)
            except UserError as e:
                errors.append(e.args[0] if e.args else str(e))
                continue

            all_rows.append(row_norm)
            key = (tmpl_code, code, rtype, str(header_qty))
            groups.setdefault(key, []).append(row_norm)

        if errors:
            msg = _(
                "The file contains errors. Please fix them and try again:\n- "
            ) + "\n- ".join(errors[:25])
            if len(errors) > 25:
                msg += _("\n...and %s more.") % (len(errors) - 25)
            raise UserError(msg)

        # Duplicate BoM prevention
        Bom = self.env["mrp.bom"].with_context(active_test=False)
        dup_errors = []
        for (tmpl_code, code, rtype, header_qty_str), rows in groups.items():
            tmpl = (
                self.env["product.template"]
                .with_context(active_test=False)
                .search([("default_code", "=", tmpl_code)], limit=1)
            )
            if tmpl:
                existing = Bom.search(
                    [
                        ("product_tmpl_id", "=", tmpl.id),
                        ("code", "=", code),
                        ("type", "=", rtype),
                        ("product_qty", "=", float(header_qty_str)),
                        ("company_id", "=", self.company_id.id),
                    ],
                    limit=1,
                )
                if existing:
                    dup_errors.append(
                        _(
                            "Duplicate BoM detected for template '%s' with code '%s' (type=%s, qty=%s). "
                            "Please change 'code' or remove the existing BoM before importing."
                        )
                        % (tmpl_code, code, rtype, header_qty_str)
                    )
        if dup_errors:
            raise UserError(" / ".join(dup_errors))

        # Dry run (validate only)
        if self.dry_run:
            total_boms = len(groups)
            total_lines = len(all_rows)
            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": _("Validation successful"),
                    "message": _(
                        "%s BoM(s), %s line(s). No data was created (dry run)."
                    )
                    % (total_boms, total_lines),
                    "sticky": False,
                },
            }

        # Create records
        created_boms = self.env["mrp.bom"]
        for (tmpl_code, code, rtype, header_qty_str), rows in groups.items():
            hdr_name = next((r["tmpl_name"] for r in rows if r.get("tmpl_name")), None)
            hdr_uom_id = next(
                (
                    r["header_uom_id"]
                    for r in rows
                    if r.get("header_uom_id") is not None
                ),
                None,
            )

            hdr_standard_price = next(
                (r["standard_price"] for r in rows if r.get("standard_price") is not None),
                None,
            )

            # Ensure uom exists (ID mode, required)
            self._get_uom_by_id(hdr_uom_id)

            tmpl = self._get_product_template(
                tmpl_code, name=hdr_name, uom_id=hdr_uom_id, standard_price=hdr_standard_price
            )
            header_qty = float(header_qty_str)

            bom_vals = {
                "product_tmpl_id": tmpl.id,
                "product_qty": header_qty,
                "code": code,
                "type": rtype,
                "company_id": self.company_id.id,
            }
            bom_vals = self.mutate_bom_create_vals(bom_vals, rows) or bom_vals
            bom = self.env["mrp.bom"].create(bom_vals)

            line_vals = []
            for r in rows:
                self._get_uom_by_id(r["line_product_uom_id"])

                comp = self._get_product(
                    r["line_code"],
                    name=r.get("line_name") or None,
                    uom_id=r.get("line_product_uom_id"),
                    standard_price=r.get("standard_price"),
                )
                uom_id = r["line_uom"].id
                if r["line_uom"].category_id != comp.uom_id.category_id:
                    raise UserError(
                        _(
                            "Row %s: Unit of Measure '%s' is not in the same category as component '%s'."
                        )
                        % (r["rownum"], r["line_uom"].name, r["line_code"])
                    )

                vals = {
                    "product_id": comp.id,
                    "product_qty": r["line_qty"],
                    "product_uom_id": uom_id,
                }
                vals = self.mutate_line_create_vals(vals, r) or vals
                line_vals.append((0, 0, vals))

            if line_vals:
                bom.write({"bom_line_ids": line_vals})
            created_boms |= bom

        # Return BoMs in view
        action = self.env.ref("mrp.mrp_bom_form_action").read()[0]
        action["domain"] = [("id", "in", created_boms.ids)]
        action["context"] = dict(
            self.env.context, default_company_id=self.company_id.id
        )
        return action
